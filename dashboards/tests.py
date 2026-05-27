from datetime import date
from io import BytesIO
from unittest.mock import patch

from django.core.cache import cache
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .models import (
    Categoria,
    ClienteFornecedor,
    ContaPagar,
    ContaReceber,
    ContaDRE,
    DeParaCategoriaDRE,
    DREProjetado,
    IndicadorConfiguracao,
    ParametroEmpresa,
)
from .forms import ContaDREForm
from .omie import importar_clientes_fornecedores_omie
from .omie import listar_contas_correntes_omie
from .views import _construir_linhas_dre, _montar_opcoes_periodo, _resolver_meses_periodo


class PeriodoDashboardTestCase(TestCase):
    def test_periodo_usa_janela_de_cinco_anos_para_frente_e_para_tras(self):
        ano_atual = date.today().year
        opcoes = _montar_opcoes_periodo('data_emissao')
        anos = [opcao['ano'] for opcao in opcoes if opcao.get('nivel') == 'ano']

        self.assertEqual(anos[0], ano_atual - 5)
        self.assertEqual(anos[-1], ano_atual + 5)
        self.assertNotIn(2039, anos)
        self.assertNotIn(2040, anos)

    def test_periodo_mensal_resolve_apenas_o_mes_escolhido(self):
        meses = _resolver_meses_periodo('mes:2026:04')

        self.assertEqual(meses, [{
            'codigo': '2026-04',
            'ano': 2026,
            'mes': 4,
            'rotulo': 'Abril',
        }])


class DreHierarquiaTestCase(TestCase):
    def test_dre_monta_cliente_fornecedor_abaixo_da_categoria(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-teste',
            nome_empresa='Empresa Teste',
        )
        pai = ContaDRE.objects.create(
            nome='Receita Bruta',
            nivel=1,
            sinal='positivo',
            ordem=1,
            ativo=True,
        )
        filho = ContaDRE.objects.create(
            nome='Receita de Servicos',
            pai=pai,
            nivel=2,
            sinal='positivo',
            ordem=1,
            ativo=True,
        )
        categoria = Categoria.objects.create(
            codigo='1.01.01',
            descricao='Mensalidades',
        )
        DeParaCategoriaDRE.objects.create(
            empresa=empresa,
            categoria=categoria,
            conta_dre=filho,
        )
        ClienteFornecedor.objects.create(
            codigo_omie=1001,
            razao_social='Cliente Um LTDA',
            nome_fantasia='Cliente Um',
        )
        ContaReceber.objects.create(
            codigo_lancamento_omie=1,
            codigo_cliente_fornecedor=1001,
            codigo_categoria='1.01.01',
            data_emissao=date(2026, 1, 10),
            data_vencimento=date(2026, 1, 20),
            valor_documento=1500,
        )
        ContaReceber.objects.create(
            codigo_lancamento_omie=2,
            codigo_cliente_fornecedor=1002,
            codigo_categoria='1.01.01',
            data_emissao=date(2026, 1, 15),
            data_vencimento=date(2026, 1, 25),
            valor_documento=500,
        )

        meses_periodo = [
            {'codigo': '2026-01', 'ano': 2026, 'mes': 1, 'rotulo': 'Jan/26'},
            {'codigo': '2026-02', 'ano': 2026, 'mes': 2, 'rotulo': 'Fev/26'},
        ]

        dre = _construir_linhas_dre(empresa, meses_periodo, 'data_emissao')

        self.assertEqual([linha['nivel'] for linha in dre], [1, 2, 3, 4, 4])
        self.assertEqual(dre[2]['nome'], 'Mensalidades')
        self.assertTrue(dre[2]['possui_filhos'])
        self.assertEqual(dre[3]['parent_id'], dre[2]['id'])
        self.assertEqual(dre[4]['parent_id'], dre[2]['id'])
        self.assertEqual(dre[3]['nome'], 'Cliente Um')
        self.assertEqual(dre[4]['nome'], 'Cliente/Fornecedor 1002')
        self.assertEqual(dre[2]['meses'][0]['realizado'], 'R$ 2.000,00')
        self.assertEqual(dre[3]['meses'][0]['realizado'], 'R$ 1.500,00')
        self.assertEqual(dre[4]['meses'][0]['realizado'], 'R$ 500,00')


class ContaDREFormTestCase(TestCase):
    def test_formulario_salva_flag_kpi(self):
        form = ContaDREForm(data={
            'nome': 'Margem EBITDA',
            'pai': '',
            'sinal': 'resultado',
            'ordem': 10,
            'kpi': 'on',
            'ativo': 'on',
        })

        self.assertTrue(form.is_valid(), form.errors)

        conta = form.save()

        self.assertTrue(conta.kpi)


class IndicadoresViewTestCase(TestCase):
    def test_view_lista_apenas_contas_kpi(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-indicadores',
            nome_empresa='Empresa Indicadores',
        )
        conta_kpi = ContaDRE.objects.create(
            nome='Receita KPI',
            nivel=1,
            sinal='positivo',
            ordem=1,
            kpi=True,
            ativo=True,
        )
        ContaDRE.objects.create(
            nome='Conta Comum',
            nivel=1,
            sinal='positivo',
            ordem=2,
            kpi=False,
            ativo=True,
        )

        response = self.client.get(reverse('indicadores', kwargs={'slug': empresa.slug_empresa}))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'indicadores.html')
        self.assertEqual(list(response.context['contas_kpi']), [conta_kpi])

    def test_view_salva_configuracao_dos_indicadores(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-indicadores',
            nome_empresa='Empresa Indicadores',
        )
        conta_receita = ContaDRE.objects.create(
            nome='Receita Bruta',
            nivel=1,
            sinal='positivo',
            ordem=1,
            kpi=True,
            ativo=True,
        )
        conta_deducao = ContaDRE.objects.create(
            nome='Impostos',
            nivel=1,
            sinal='negativo',
            ordem=2,
            kpi=True,
            ativo=True,
        )

        response = self.client.post(
            reverse('indicadores', kwargs={'slug': empresa.slug_empresa}),
            data={
                'receita_bruta__conta_1': str(conta_receita.id),
                'deducoes_gastos_variaveis__conta_1': str(conta_deducao.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            IndicadorConfiguracao.objects.filter(
                empresa=empresa,
                chave='receita_bruta',
                conta_1=conta_receita,
            ).exists()
        )
        self.assertTrue(
            IndicadorConfiguracao.objects.filter(
                empresa=empresa,
                chave='deducoes_gastos_variaveis',
                conta_1=conta_deducao,
            ).exists()
        )

    def test_view_salva_quatro_contas_na_margem_de_contribuicao(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-margem',
            nome_empresa='Empresa Margem',
        )
        conta_1 = ContaDRE.objects.create(
            nome='Receita Liquida',
            nivel=1,
            sinal='positivo',
            ordem=1,
            kpi=True,
            ativo=True,
        )
        conta_2 = ContaDRE.objects.create(
            nome='Custos Variaveis',
            nivel=1,
            sinal='negativo',
            ordem=2,
            kpi=True,
            ativo=True,
        )
        conta_3 = ContaDRE.objects.create(
            nome='Impostos',
            nivel=1,
            sinal='negativo',
            ordem=3,
            kpi=True,
            ativo=True,
        )
        conta_4 = ContaDRE.objects.create(
            nome='Receita Bruta',
            nivel=1,
            sinal='positivo',
            ordem=4,
            kpi=True,
            ativo=True,
        )

        response = self.client.post(
            reverse('indicadores', kwargs={'slug': empresa.slug_empresa}),
            data={
                'margem_contribuicao__conta_1': str(conta_1.id),
                'margem_contribuicao__conta_2': str(conta_2.id),
                'margem_contribuicao__conta_3': str(conta_3.id),
                'margem_contribuicao__conta_4': str(conta_4.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            IndicadorConfiguracao.objects.filter(
                empresa=empresa,
                chave='margem_contribuicao',
                conta_1=conta_1,
                conta_2=conta_2,
                conta_3=conta_3,
                conta_4=conta_4,
            ).exists()
        )


class DashboardIndicadoresTestCase(TestCase):
    def test_dashboard_dre_projetado_renderiza_layout_com_projetado(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-projetado',
            nome_empresa='Empresa Projetado',
        )
        receita_pai = ContaDRE.objects.create(nome='Receita Bruta', nivel=1, sinal='positivo', ordem=1, ativo=True)
        receita_filho = ContaDRE.objects.create(nome='Receita Operacional', pai=receita_pai, nivel=2, sinal='positivo', ordem=1, ativo=True)
        DREProjetado.objects.create(
            empresa=empresa,
            tipo_linha='conta',
            conta_dre=receita_pai,
            ano=2026,
            mes=1,
            valor=1000,
        )

        response = self.client.get(
            reverse('dashboard_dre_projetado', kwargs={'slug': empresa.slug_empresa}),
            {'periodo': 'ano:2026'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'dashboard2.html')
        self.assertContains(response, 'DRE vs Projetado')
        self.assertContains(response, 'Projetado')
        self.assertContains(response, 'Realizado')
        self.assertContains(response, 'dreProjetadoContaSelect')
        self.assertContains(response, 'Projetado vs Realizado')
        self.assertEqual(response.context['grafico_projetado_realizado']['opcoes'][0]['nome'], 'Receita Bruta')
        self.assertEqual(response.context['grafico_projetado_realizado']['opcoes'][0]['filhos'][0]['nome'], 'Receita Operacional')
        self.assertEqual(response.context['grafico_projetado_realizado']['series'][0]['pontos'][0]['projetado'], 1000.0)

    def test_exportar_planilha_dre_projetado_traz_contas_e_categorias_sem_clientes(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-projetado-export',
            nome_empresa='Empresa Projetado Export',
        )
        receita_pai = ContaDRE.objects.create(nome='Receita Bruta', nivel=1, sinal='positivo', ordem=1, ativo=True)
        receita_filho = ContaDRE.objects.create(nome='Receita Operacional', pai=receita_pai, nivel=2, sinal='positivo', ordem=1, ativo=True)
        categoria = Categoria.objects.create(codigo='1.01.01', descricao='Receita Servicos')
        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria, conta_dre=receita_filho)
        DREProjetado.objects.create(
            empresa=empresa,
            tipo_linha='categoria',
            conta_dre=receita_filho,
            categoria=categoria,
            ano=2026,
            mes=1,
            valor=1234,
        )

        response = self.client.get(
            reverse('exportar_planilha_dre_projetado', kwargs={'slug': empresa.slug_empresa}),
            {'periodo': 'ano:2026'},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        worksheet = workbook.active
        self.assertEqual(worksheet['E1'].value, 'Janeiro')
        nomes = [worksheet.cell(row=linha, column=4).value for linha in range(2, worksheet.max_row + 1)]
        self.assertIn('(+) Receita Bruta', nomes)
        self.assertIn('(+) Receita Operacional', nomes)
        self.assertIn('Receita Servicos', [nome.strip() for nome in nomes])
        self.assertEqual(worksheet.cell(row=4, column=5).value, 1234)
        self.assertNotIn('Sem cliente/fornecedor', nomes)

    def test_importar_planilha_dre_projetado_salva_valores_no_painel(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-projetado-import',
            nome_empresa='Empresa Projetado Import',
        )
        receita = ContaDRE.objects.create(nome='Receita Bruta', nivel=1, sinal='positivo', ordem=1, ativo=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Tipo', 'Conta DRE ID', 'Categoria ID', 'Conta / Categoria', 'Janeiro', 'Fevereiro'])
        worksheet.append(['conta', receita.id, '', 'Receita Bruta', 1000, 1500])
        arquivo = BytesIO()
        workbook.save(arquivo)
        arquivo.seek(0)
        arquivo.name = 'dre_projetado.xlsx'

        response = self.client.post(
            reverse('importar_planilha_dre_projetado', kwargs={'slug': empresa.slug_empresa}),
            data={
                'ano_projetado': '2026',
                'planilha_projetado': arquivo,
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(DREProjetado.objects.filter(
            empresa=empresa,
            conta_dre=receita,
            ano=2026,
            mes=1,
            valor=1000,
        ).exists())

        painel = self.client.get(
            reverse('dashboard_dre_projetado', kwargs={'slug': empresa.slug_empresa}),
            {'periodo': 'ano:2026'},
        )
        self.assertContains(painel, 'R$ 1.000,00')

    def test_dashboard_resultado_exibe_kpis_calculados_a_partir_do_dre(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-kpi',
            nome_empresa='Empresa KPI',
        )

        receita_pai = ContaDRE.objects.create(nome='Receita Bruta', nivel=1, sinal='positivo', ordem=1, kpi=True, ativo=True)
        receita_filho = ContaDRE.objects.create(nome='Receita Operacional', pai=receita_pai, nivel=2, sinal='positivo', ordem=1, kpi=True, ativo=True)
        deducoes_pai = ContaDRE.objects.create(nome='Deducoes e Gastos Variaveis', nivel=1, sinal='negativo', ordem=2, ativo=True)
        imposto = ContaDRE.objects.create(nome='Impostos', pai=deducoes_pai, nivel=2, sinal='negativo', ordem=1, kpi=True, ativo=True)
        comissao = ContaDRE.objects.create(nome='Comissoes', pai=deducoes_pai, nivel=2, sinal='negativo', ordem=2, kpi=True, ativo=True)
        margem = ContaDRE.objects.create(nome='Margem de Contribuicao', nivel=1, sinal='resultado', ordem=3, kpi=True, ativo=True)
        fixos_pai = ContaDRE.objects.create(nome='Gastos Fixos', nivel=1, sinal='negativo', ordem=4, ativo=True)
        administrativo = ContaDRE.objects.create(nome='Administrativo', pai=fixos_pai, nivel=2, sinal='negativo', ordem=1, kpi=True, ativo=True)
        ebit = ContaDRE.objects.create(nome='Resultado do Exercicio', nivel=1, sinal='resultado', ordem=5, kpi=True, ativo=True)

        categoria_receita = Categoria.objects.create(codigo='1.01.01', descricao='Receita')
        categoria_imposto = Categoria.objects.create(codigo='2.01.01', descricao='Impostos')
        categoria_comissao = Categoria.objects.create(codigo='2.01.02', descricao='Comissao')
        categoria_fixo = Categoria.objects.create(codigo='2.02.01', descricao='Administrativo')

        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_receita, conta_dre=receita_filho)
        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_imposto, conta_dre=imposto)
        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_comissao, conta_dre=comissao)
        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_fixo, conta_dre=administrativo)

        ContaReceber.objects.create(
            codigo_lancamento_omie=10,
            codigo_categoria='1.01.01',
            data_emissao=date(2026, 1, 5),
            data_vencimento=date(2026, 1, 5),
            valor_documento=1000,
        )
        ContaPagar.objects.create(
            codigo_lancamento_omie=20,
            codigo_categoria='2.01.01',
            data_emissao=date(2026, 1, 6),
            data_vencimento=date(2026, 1, 6),
            valor_documento=200,
        )
        ContaPagar.objects.create(
            codigo_lancamento_omie=21,
            codigo_categoria='2.01.02',
            data_emissao=date(2026, 1, 7),
            data_vencimento=date(2026, 1, 7),
            valor_documento=100,
        )
        ContaPagar.objects.create(
            codigo_lancamento_omie=22,
            codigo_categoria='2.02.01',
            data_emissao=date(2026, 1, 8),
            data_vencimento=date(2026, 1, 8),
            valor_documento=300,
        )

        IndicadorConfiguracao.objects.create(empresa=empresa, chave='receita_bruta', conta_1=receita_pai)
        IndicadorConfiguracao.objects.create(
            empresa=empresa,
            chave='deducoes_gastos_variaveis',
            conta_1=imposto,
            conta_2=comissao,
        )
        IndicadorConfiguracao.objects.create(
            empresa=empresa,
            chave='margem_contribuicao',
            conta_1=margem,
            conta_2=imposto,
            conta_3=comissao,
            conta_4=receita_pai,
        )
        IndicadorConfiguracao.objects.create(empresa=empresa, chave='gastos_fixos', conta_1=administrativo)
        IndicadorConfiguracao.objects.create(empresa=empresa, chave='ebit', conta_1=ebit)

        response = self.client.get(
            reverse('dashboard_resultado', kwargs={'slug': empresa.slug_empresa}),
            data={'periodo': 'ano:2026'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['kpis']['receita_bruta'], 'R$ 1,00 mil')
        self.assertEqual(response.context['kpis']['deducoes_gastos_variaveis'], '-R$ 300,00')
        self.assertEqual(response.context['kpis']['margem_contribuicao'], '40,00%')
        self.assertEqual(response.context['kpis']['gastos_fixos'], '-R$ 300,00')
        self.assertEqual(response.context['kpis']['ebit'], '40,00%')

    def test_dashboard_resultado_destaca_valores_negativos_em_vermelho(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-kpi-negativo',
            nome_empresa='Empresa KPI Negativo',
        )

        receita_pai = ContaDRE.objects.create(nome='Receita Bruta', nivel=1, sinal='positivo', ordem=1, kpi=True, ativo=True)
        receita_filho = ContaDRE.objects.create(nome='Receita Operacional', pai=receita_pai, nivel=2, sinal='positivo', ordem=1, kpi=True, ativo=True)
        deducoes_pai = ContaDRE.objects.create(nome='Deducoes e Gastos Variaveis', nivel=1, sinal='negativo', ordem=2, ativo=True)
        imposto = ContaDRE.objects.create(nome='Impostos', pai=deducoes_pai, nivel=2, sinal='negativo', ordem=1, kpi=True, ativo=True)

        categoria_receita = Categoria.objects.create(codigo='1.01.01', descricao='Receita')
        categoria_imposto = Categoria.objects.create(codigo='2.01.01', descricao='Impostos')

        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_receita, conta_dre=receita_filho)
        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_imposto, conta_dre=imposto)

        ContaReceber.objects.create(
            codigo_lancamento_omie=30,
            codigo_categoria='1.01.01',
            data_emissao=date(2026, 1, 5),
            data_vencimento=date(2026, 1, 5),
            valor_documento=1000,
        )
        ContaPagar.objects.create(
            codigo_lancamento_omie=31,
            codigo_categoria='2.01.01',
            data_emissao=date(2026, 1, 6),
            data_vencimento=date(2026, 1, 6),
            valor_documento=200,
        )

        IndicadorConfiguracao.objects.create(
            empresa=empresa,
            chave='deducoes_gastos_variaveis',
            conta_1=imposto,
        )

        response = self.client.get(
            reverse('dashboard_resultado', kwargs={'slug': empresa.slug_empresa}),
            data={'periodo': 'ano:2026'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<h5 class="valor-negativo">-R$ 200,00</h5>', html=True)
        self.assertContains(response, '<td class="valor-negativo">-R$ 200,00</td>', html=True)

    def test_dashboard_resultado_calcula_ah_do_primeiro_mes_com_mes_anterior(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-ah-primeiro-mes',
            nome_empresa='Empresa AH Primeiro Mes',
        )

        receita_pai = ContaDRE.objects.create(nome='Receita Bruta', nivel=1, sinal='positivo', ordem=1, ativo=True)
        receita_filho = ContaDRE.objects.create(nome='Receita Operacional', pai=receita_pai, nivel=2, sinal='positivo', ordem=1, ativo=True)
        categoria_receita = Categoria.objects.create(codigo='1.01.01', descricao='Receita')

        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_receita, conta_dre=receita_filho)

        ContaReceber.objects.create(
            codigo_lancamento_omie=32,
            codigo_categoria='1.01.01',
            data_emissao=date(2025, 12, 10),
            data_vencimento=date(2025, 12, 10),
            valor_documento=500,
        )
        ContaReceber.objects.create(
            codigo_lancamento_omie=33,
            codigo_categoria='1.01.01',
            data_emissao=date(2026, 1, 10),
            data_vencimento=date(2026, 1, 10),
            valor_documento=1000,
        )

        response = self.client.get(
            reverse('dashboard_resultado', kwargs={'slug': empresa.slug_empresa}),
            data={
                'periodo': 'customizado',
                'data_inicial': '2026-01-01',
                'data_final': '2026-05-31',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['meses_dre']), 5)
        self.assertEqual(response.context['dre'][0]['meses'][0]['realizado'], 'R$ 1.000,00')
        self.assertEqual(response.context['dre'][0]['meses'][0]['ah'], '100,0%')
        self.assertEqual(response.context['dre'][0]['meses'][0]['ah_direcao'], 'alta')

    def test_dashboard_resultado_exibe_grafico_para_contas_pai_do_dre(self):
        empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-grafico-dre',
            nome_empresa='Empresa Grafico DRE',
        )

        receita_pai = ContaDRE.objects.create(nome='Receita Bruta Operacional', nivel=1, sinal='positivo', ordem=1, ativo=True)
        receita_filho = ContaDRE.objects.create(nome='Receita Operacional', pai=receita_pai, nivel=2, sinal='positivo', ordem=1, ativo=True)
        gastos_pai = ContaDRE.objects.create(nome='Gastos Fixos', nivel=1, sinal='negativo', ordem=2, ativo=True)
        gastos_filho = ContaDRE.objects.create(nome='Administrativo', pai=gastos_pai, nivel=2, sinal='negativo', ordem=1, ativo=True)

        categoria_receita = Categoria.objects.create(codigo='1.01.01', descricao='Receita')
        categoria_gastos = Categoria.objects.create(codigo='2.01.01', descricao='Gastos')
        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_receita, conta_dre=receita_filho)
        DeParaCategoriaDRE.objects.create(empresa=empresa, categoria=categoria_gastos, conta_dre=gastos_filho)

        ContaReceber.objects.create(
            codigo_lancamento_omie=40,
            codigo_categoria='1.01.01',
            data_emissao=date(2026, 1, 5),
            data_vencimento=date(2026, 1, 5),
            valor_documento=400,
        )
        ContaReceber.objects.create(
            codigo_lancamento_omie=41,
            codigo_categoria='1.01.01',
            data_emissao=date(2026, 2, 5),
            data_vencimento=date(2026, 2, 5),
            valor_documento=500,
        )
        ContaPagar.objects.create(
            codigo_lancamento_omie=42,
            codigo_categoria='2.01.01',
            data_emissao=date(2026, 1, 8),
            data_vencimento=date(2026, 1, 8),
            valor_documento=100,
        )
        ContaPagar.objects.create(
            codigo_lancamento_omie=43,
            codigo_categoria='2.01.01',
            data_emissao=date(2026, 2, 8),
            data_vencimento=date(2026, 2, 8),
            valor_documento=200,
        )

        response = self.client.get(
            reverse('dashboard_resultado', kwargs={'slug': empresa.slug_empresa}),
            data={'periodo': 'tri:2026:T1'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'dreContaPaiSelect')
        self.assertContains(response, 'Receita Bruta Operacional')
        self.assertTrue(response.context['grafico_contas_pai_dre'])
        self.assertEqual(response.context['grafico_contas_pai_dre'][0]['colunas'][0]['rotulo'], 'Jan')
        self.assertEqual(response.context['grafico_contas_pai_dre'][0]['colunas'][1]['rotulo'], 'Fev')
        self.assertEqual(response.context['grafico_contas_pai_dre'][0]['colunas'][0]['valor_formatado'], 'R$ 400,00')
        self.assertEqual(response.context['grafico_variacao_mensal']['conta'], 'Receita Bruta Operacional')
        self.assertEqual(response.context['grafico_variacao_mensal']['colunas'][0]['rotulo'], 'Jan')
        self.assertEqual(response.context['grafico_variacao_mensal']['colunas'][1]['rotulo'], 'Fev')
        self.assertEqual(response.context['grafico_variacao_mensal']['colunas'][1]['percentual'], '+25%')
        self.assertEqual(response.context['grafico_variacao_mensal']['colunas'][1]['direcao'], 'alta')

        gastos_grafico = response.context['grafico_contas_pai_dre'][1]
        self.assertEqual(gastos_grafico['nome'], 'Gastos Fixos')
        self.assertEqual(gastos_grafico['colunas'][0]['valor_formatado'], '-R$ 100,00')
        self.assertEqual(gastos_grafico['variacao_mensal']['conta'], 'Gastos Fixos')
        self.assertEqual(gastos_grafico['variacao_mensal']['colunas'][1]['percentual'], '+100%')
        self.assertEqual(gastos_grafico['variacao_mensal']['colunas'][1]['direcao'], 'alta')


class ParametrosOmieViewTestCase(TestCase):
    def setUp(self):
        self.empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-omie',
            nome_empresa='Empresa Omie',
        )
        self.url = reverse('parametros_gerais', kwargs={'slug': self.empresa.slug_empresa})

    def test_parametros_empresa_exibe_grupos_de_parametros(self):
        response = self.client.get(
            reverse('parametros_empresa', kwargs={'slug': self.empresa.slug_empresa})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Parâmetros gerais')
        self.assertContains(response, 'Dashboard Resultado')

    @patch('dashboards.views.sincronizar_dados_omie')
    def test_salvar_credenciais_nao_dispara_sincronizacao(self, sincronizar_mock):
        response = self.client.post(
            self.url,
            data={
                'acao': 'salvar_credenciais',
                'app_key_omie': 'nova-key',
                'app_secret_omie': 'novo-secret',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.empresa.refresh_from_db()
        self.assertEqual(self.empresa.app_key_omie, 'nova-key')
        self.assertEqual(self.empresa.app_secret_omie, 'novo-secret')
        sincronizar_mock.assert_not_called()
        self.assertEqual(response.context['modal_tipo'], 'success')

    @patch('dashboards.views.sincronizar_dados_omie')
    def test_sincronizacao_usa_credenciais_salvas(self, sincronizar_mock):
        self.empresa.app_key_omie = 'key-salva'
        self.empresa.app_secret_omie = 'secret-salvo'
        self.empresa.save(update_fields=['app_key_omie', 'app_secret_omie'])
        sincronizar_mock.return_value = {
            'categorias': 10,
            'clientes_fornecedores': 20,
            'contas_correntes': 3,
            'contas_pagar': 40,
            'contas_receber': 50,
        }

        response = self.client.post(
            self.url,
            data={'acao': 'sincronizar_omie'},
        )

        self.assertEqual(response.status_code, 200)
        sincronizar_mock.assert_called_once_with('key-salva', 'secret-salvo')
        self.assertEqual(response.context['modal_tipo'], 'success')
        self.assertIn('10 categorias', response.context['modal_mensagem'])

    @patch('dashboards.views.sincronizar_dados_omie')
    def test_sincronizacao_sem_credenciais_exibe_erro(self, sincronizar_mock):
        response = self.client.post(
            self.url,
            data={'acao': 'sincronizar_omie'},
        )

        self.assertEqual(response.status_code, 200)
        sincronizar_mock.assert_not_called()
        self.assertEqual(response.context['modal_tipo'], 'error')
        self.assertIn('Salve a App Key e a App Secret', response.context['modal_mensagem'])

    @patch('dashboards.views._iniciar_sincronizacao_omie')
    def test_endpoint_inicia_sincronizacao_assincrona(self, iniciar_mock):
        self.empresa.app_key_omie = 'key-salva'
        self.empresa.app_secret_omie = 'secret-salvo'
        self.empresa.save(update_fields=['app_key_omie', 'app_secret_omie'])
        iniciar_mock.return_value = 'task-123'

        response = self.client.post(
            reverse('iniciar_sincronizacao_omie', kwargs={'slug': self.empresa.slug_empresa}),
            content_type='application/json',
            data='{}',
        )

        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            response.content,
            {
                'ok': True,
                'task_id': 'task-123',
                'mensagem': 'Sincronizacao iniciada com sucesso.',
            },
        )
        iniciar_mock.assert_called_once_with(self.empresa)

    def test_endpoint_status_retorna_progresso(self):
        cache.set('omie_sync_status:task-123', {
            'empresa_id': self.empresa.id,
            'task_id': 'task-123',
            'state': 'running',
            'percentual': 47,
            'etapa': 'clientes_fornecedores',
            'fase': 'gravando',
            'mensagem': 'Gravando clientes e fornecedores no banco... (47/100)',
        }, 300)

        response = self.client.get(
            reverse('status_sincronizacao_omie', kwargs={'slug': self.empresa.slug_empresa, 'task_id': 'task-123'})
        )

        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            response.content,
            {
                'ok': True,
                'empresa_id': self.empresa.id,
                'task_id': 'task-123',
                'state': 'running',
                'percentual': 47,
                'etapa': 'clientes_fornecedores',
                'fase': 'gravando',
                'mensagem': 'Gravando clientes e fornecedores no banco... (47/100)',
            },
        )


class OmieClientesFornecedoresImportacaoTestCase(TestCase):
    @patch('dashboards.omie._listar_registros_paginados')
    def test_importacao_clientes_fornecedores_faz_fallback_para_listagem_resumida(self, listar_mock):
        listar_mock.side_effect = [
            [],
            [{
                'codigo_cliente_omie': '101',
                'codigo_cliente_integracao': 'CF-101',
                'razao_social': 'Fornecedor XPTO LTDA',
                'nome_fantasia': 'Fornecedor XPTO',
                'cnpj_cpf': '12.345.678/0001-99',
            }],
        ]

        total = importar_clientes_fornecedores_omie('app-key', 'app-secret')

        self.assertEqual(total, 1)
        self.assertTrue(
            ClienteFornecedor.objects.filter(
                codigo_omie=101,
                codigo_integracao='CF-101',
                razao_social='Fornecedor XPTO LTDA',
                nome_fantasia='Fornecedor XPTO',
            ).exists()
        )

    @patch('dashboards.omie._listar_registros_paginados')
    def test_importacao_clientes_fornecedores_usa_listagem_completa_como_padrao(self, listar_mock):
        listar_mock.return_value = [{
            'codigo_cliente_omie': '202',
            'codigo_cliente_integracao': 'CF-202',
            'razao_social': 'Cliente ABC LTDA',
            'nome_fantasia': 'Cliente ABC',
            'cnpj_cpf': '11.222.333/0001-44',
        }]

        total = importar_clientes_fornecedores_omie('app-key', 'app-secret')

        self.assertEqual(total, 1)
        self.assertEqual(listar_mock.call_args.kwargs['call'], 'ListarClientes')
        self.assertIn('clientes_cadastro', listar_mock.call_args.kwargs['response_keys'])
        self.assertTrue(ClienteFornecedor.objects.filter(codigo_omie=202).exists())


class OmieContasCorrentesListagemTestCase(TestCase):
    @patch('dashboards.omie._listar_registros_paginados')
    def test_listagem_contas_correntes_aceita_chave_oficial_da_omie(self, listar_mock):
        listar_mock.return_value = [
            {'nCodCC': 2, 'descricao': 'Banco B'},
            {'nCodCC': 1, 'descricao': 'Banco A'},
        ]

        contas = listar_contas_correntes_omie('app-key', 'app-secret')

        self.assertEqual([conta['nCodCC'] for conta in contas], [1, 2])
        self.assertEqual(listar_mock.call_args.kwargs['call'], 'ListarContasCorrentes')
        self.assertIn('ListarContasCorrentes', listar_mock.call_args.kwargs['response_keys'])
        self.assertEqual(listar_mock.call_args.kwargs['parametros_extras'], {'apenas_importado_api': 'N'})


class DeParaPlanilhaTestCase(TestCase):
    def setUp(self):
        self.empresa = ParametroEmpresa.objects.create(
            slug_empresa='empresa-depara',
            nome_empresa='Empresa DePara',
        )
        self.conta_receita = ContaDRE.objects.create(
            nome='Receita Operacional',
            nivel=2,
            sinal='positivo',
            ordem=1,
            ativo=True,
        )
        self.conta_despesa = ContaDRE.objects.create(
            nome='Custos Variaveis',
            nivel=2,
            sinal='negativo',
            ordem=2,
            ativo=True,
        )
        self.categoria_pai = Categoria.objects.create(codigo='1.01', descricao='Receitas Diretas')
        self.categoria_filha = Categoria.objects.create(codigo='1.01.01', descricao='Venda de Produtos')
        self.categoria_sem_preenchimento = Categoria.objects.create(codigo='1.01.02', descricao='Venda de Servicos')

    def _criar_planilha(self, linhas):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Codigo da Categoria', 'Nome da Categoria', 'Conta DRE'])
        for linha in linhas:
            worksheet.append(linha)

        arquivo = BytesIO()
        workbook.save(arquivo)
        arquivo.seek(0)
        arquivo.name = 'depara.xlsx'
        return arquivo

    def test_exportar_planilha_depara(self):
        DeParaCategoriaDRE.objects.create(
            empresa=self.empresa,
            categoria=self.categoria_filha,
            conta_dre=self.conta_receita,
        )

        response = self.client.get(reverse('exportar_planilha_depara', kwargs={'slug': self.empresa.slug_empresa}))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet['A1'].value, 'Codigo da Categoria')
        self.assertEqual(worksheet['B2'].value, 'Receitas Diretas')
        self.assertEqual(worksheet['C2'].value, 'Conta totalizadora')
        self.assertTrue(worksheet['A2'].font.bold)
        self.assertTrue(worksheet['B2'].font.bold)
        self.assertTrue(worksheet['C2'].font.bold)
        self.assertEqual(worksheet['A3'].value, '1.01.01')
        self.assertEqual(worksheet['C3'].value, 'Receita Operacional')

    def test_importar_planilha_depara_sobrepoe_mapeamentos_validos(self):
        DeParaCategoriaDRE.objects.create(
            empresa=self.empresa,
            categoria=self.categoria_filha,
            conta_dre=self.conta_receita,
        )
        planilha = self._criar_planilha([
            ['1.01', 'Receitas Diretas', 'Custos Variaveis'],
            ['1.01.01', 'Venda de Produtos', 'Custos Variaveis'],
            ['1.01.02', 'Venda de Servicos', ''],
        ])

        response = self.client.post(
            reverse('depara_categoria_dre', kwargs={'slug': self.empresa.slug_empresa}),
            data={'acao': 'importar_planilha', 'planilha_depara': planilha},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Planilha importada com sucesso.')
        self.assertEqual(
            DeParaCategoriaDRE.objects.get(empresa=self.empresa, categoria=self.categoria_filha).conta_dre,
            self.conta_despesa,
        )
        self.assertFalse(
            DeParaCategoriaDRE.objects.filter(
                empresa=self.empresa,
                categoria=self.categoria_sem_preenchimento,
            ).exists()
        )

    def test_importar_planilha_invalida_exibe_modal_e_nao_importa(self):
        planilha = self._criar_planilha([
            ['1.01.01', 'Venda de Produtos', 'Conta Inexistente'],
        ])

        response = self.client.post(
            reverse('depara_categoria_dre', kwargs={'slug': self.empresa.slug_empresa}),
            data={'acao': 'importar_planilha', 'planilha_depara': planilha},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'A planilha nao foi importada por motivos de preenchimento')
        self.assertFalse(DeParaCategoriaDRE.objects.filter(empresa=self.empresa).exists())

    def test_importar_planilha_aceita_variacoes_de_acentos_e_hifens(self):
        ContaDRE.objects.create(
            nome='Saidas Nao Operacionais',
            nivel=2,
            sinal='positivo',
            ordem=3,
            ativo=True,
        )
        planilha = self._criar_planilha([
            ['1.01.01', 'Venda de Produtos', 'Custos Variáveis'],
            ['1.01.02', 'Venda de Servicos', 'Saídas Não-Operacionais'],
        ])

        response = self.client.post(
            reverse('depara_categoria_dre', kwargs={'slug': self.empresa.slug_empresa}),
            data={'acao': 'importar_planilha', 'planilha_depara': planilha},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Planilha importada com sucesso.')
        self.assertEqual(
            DeParaCategoriaDRE.objects.get(empresa=self.empresa, categoria=self.categoria_filha).conta_dre.nome,
            'Custos Variaveis',
        )
        self.assertEqual(
            DeParaCategoriaDRE.objects.get(
                empresa=self.empresa,
                categoria=self.categoria_sem_preenchimento,
            ).conta_dre.nome,
            'Saidas Nao Operacionais',
        )
