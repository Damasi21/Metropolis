from calendar import monthrange
from datetime import date
import json
from threading import Thread
import unicodedata
from uuid import uuid4

from django.contrib import messages
from django.core.cache import cache
from django.db import transaction
from django.db import OperationalError, ProgrammingError
from django.db import close_old_connections
from django.db.models import Prefetch, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .forms import ContaDREForm, ParametroEmpresaForm
from .models import (
    Categoria,
    ClienteFornecedor,
    ComposicaoContaDRE,
    ContaPagar,
    ContaReceber,
    ContaDRE,
    DeParaCategoriaDRE,
    IndicadorConfiguracao,
    ParametroEmpresa,
)
from .omie import sincronizar_dados_omie


MESES_DRE = (
    ('jan', 'Janeiro'),
    ('fev', 'Fevereiro'),
)

MESES_LABELS = {
    1: 'Janeiro',
    2: 'Fevereiro',
    3: 'Marco',
    4: 'Abril',
    5: 'Maio',
    6: 'Junho',
    7: 'Julho',
    8: 'Agosto',
    9: 'Setembro',
    10: 'Outubro',
    11: 'Novembro',
    12: 'Dezembro',
}

MESES_ABREVIADOS = {
    1: 'Jan',
    2: 'Fev',
    3: 'Mar',
    4: 'Abr',
    5: 'Mai',
    6: 'Jun',
    7: 'Jul',
    8: 'Ago',
    9: 'Set',
    10: 'Out',
    11: 'Nov',
    12: 'Dez',
}

TRIMESTRES = {
    1: (1, 2, 3),
    2: (4, 5, 6),
    3: (7, 8, 9),
    4: (10, 11, 12),
}

TIPOS_VISUALIZACAO = {
    'competencia': {
        'label': 'Competência',
        'campo_data': 'data_emissao',
    },
    'caixa': {
        'label': 'Caixa',
        'campo_data': 'data_vencimento',
    },
}

INDICADORES_CONFIG = [
    {
        'chave': 'receita_bruta',
        'titulo': 'Receita Bruta',
        'campos': [('conta_1', 'Conta DRE')],
        'operador': None,
    },
    {
        'chave': 'deducoes_gastos_variaveis',
        'titulo': 'Deducoes e Gastos Variaveis',
        'campos': [('conta_1', 'Conta DRE'), ('conta_2', 'Conta DRE')],
        'operador': '+',
    },
    {
        'chave': 'margem_contribuicao',
        'titulo': 'Margem de Contribuicao',
        'campos': [
            ('conta_1', 'Conta base'),
            ('conta_2', 'Conta soma 1'),
            ('conta_3', 'Conta soma 2'),
            ('conta_4', 'Conta denominador'),
        ],
        'operadores': ['-', '+', '/'],
        'formula_legenda': '(conta DRE - (conta DRE + conta DRE)) / conta DRE',
    },
    {
        'chave': 'gastos_fixos',
        'titulo': 'Gastos Fixos',
        'campos': [('conta_1', 'Conta DRE')],
        'operador': None,
    },
    {
        'chave': 'ebit',
        'titulo': 'EBIT',
        'campos': [('conta_1', 'Conta DRE')],
        'operador': None,
    },
]
CHAVE_CACHE_SINCRONIZACAO_OMIE = 'omie_sync_status'
TTL_CACHE_SINCRONIZACAO_OMIE = 60 * 60


def _chave_status_sincronizacao(task_id):
    return f'{CHAVE_CACHE_SINCRONIZACAO_OMIE}:{task_id}'


def _salvar_status_sincronizacao(task_id, dados):
    cache.set(_chave_status_sincronizacao(task_id), dados, TTL_CACHE_SINCRONIZACAO_OMIE)


def _obter_status_sincronizacao(task_id):
    return cache.get(_chave_status_sincronizacao(task_id))


def _montar_mensagem_sincronizacao_concluida(totais):
    return (
        'Sincronizacao concluida com sucesso. '
        f"{totais['categorias']} categorias, "
        f"{totais['clientes_fornecedores']} clientes/fornecedores, "
        f"{totais['contas_correntes']} contas correntes, "
        f"{totais['contas_pagar']} contas a pagar e "
        f"{totais['contas_receber']} contas a receber gravados/atualizados no banco."
    )


def _executar_sincronizacao_omie_em_background(task_id, empresa_id, app_key, app_secret):
    close_old_connections()

    def atualizar_progresso(payload):
        status_atual = _obter_status_sincronizacao(task_id) or {}
        _salvar_status_sincronizacao(task_id, {
            **status_atual,
            'empresa_id': empresa_id,
            'task_id': task_id,
            'state': 'running',
            'percentual': payload.get('percentual', status_atual.get('percentual', 0)),
            'etapa': payload.get('etapa', status_atual.get('etapa')),
            'fase': payload.get('fase', status_atual.get('fase')),
            'mensagem': payload.get('mensagem', status_atual.get('mensagem')),
        })

    try:
        totais = sincronizar_dados_omie(
            app_key,
            app_secret,
            progress_callback=atualizar_progresso,
        )
        _salvar_status_sincronizacao(task_id, {
            'empresa_id': empresa_id,
            'task_id': task_id,
            'state': 'completed',
            'percentual': 100,
            'etapa': 'finalizado',
            'fase': 'concluido',
            'mensagem': _montar_mensagem_sincronizacao_concluida(totais),
            'totais': totais,
        })
    except Exception as exc:
        _salvar_status_sincronizacao(task_id, {
            'empresa_id': empresa_id,
            'task_id': task_id,
            'state': 'error',
            'percentual': 100,
            'etapa': 'erro',
            'fase': 'falha',
            'mensagem': f'Erro ao sincronizar dados da Omie: {exc}',
        })
    finally:
        close_old_connections()


def _iniciar_sincronizacao_omie(empresa):
    task_id = uuid4().hex
    _salvar_status_sincronizacao(task_id, {
        'empresa_id': empresa.id,
        'task_id': task_id,
        'state': 'running',
        'percentual': 0,
        'etapa': 'inicializando',
        'fase': 'preparando',
        'mensagem': 'Preparando sincronizacao com a Omie...',
    })
    Thread(
        target=_executar_sincronizacao_omie_em_background,
        args=(task_id, empresa.id, empresa.app_key_omie, empresa.app_secret_omie),
        daemon=True,
    ).start()
    return task_id


def _formatar_moeda(valor):
    sinal = '-' if valor < 0 else ''
    valor = abs(valor)
    texto = f'{valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f'{sinal}R$ {texto}'


def _formatar_percentual(valor):
    return f'{valor:.1f}%'.replace('.', ',')


def _formatar_percentual_grafico(valor):
    sinal = '+' if valor > 0 else ''
    return f'{sinal}{valor:.0f}%'.replace('.', ',')


def _formatar_percentual_indicador(valor):
    return f'{valor:.2f}%'.replace('.', ',')


def _formatar_moeda_resumida(valor):
    sinal = '-' if valor < 0 else ''
    valor_absoluto = abs(float(valor or 0))

    if valor_absoluto >= 1_000_000_000:
        numero = valor_absoluto / 1_000_000_000
        sufixo = ' Bi'
    elif valor_absoluto >= 1_000_000:
        numero = valor_absoluto / 1_000_000
        sufixo = ' Mi'
    elif valor_absoluto >= 1_000:
        numero = valor_absoluto / 1_000
        sufixo = ' mil'
    else:
        texto = f'{valor_absoluto:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f'{sinal}R$ {texto}'

    return f"{sinal}R$ {f'{numero:.2f}'.replace('.', ',')}{sufixo}"


def _ultimo_dia_mes(ano, mes):
    return date(ano, mes, monthrange(ano, mes)[1])


def _limites_periodo_meses(meses_periodo):
    if not meses_periodo:
        hoje = date.today()
        return date(hoje.year, hoje.month, 1), _ultimo_dia_mes(hoje.year, hoje.month)

    inicio = date(meses_periodo[0]['ano'], meses_periodo[0]['mes'], 1)
    fim = _ultimo_dia_mes(meses_periodo[-1]['ano'], meses_periodo[-1]['mes'])
    return inicio, fim


def _inicializar_valores_periodo(meses_periodo):
    return {mes['codigo']: 0.0 for mes in meses_periodo}


def _somar_valores(destino, origem):
    for codigo, valor in origem.items():
        destino[codigo] = round(destino.get(codigo, 0.0) + float(valor or 0), 2)

    return destino


def _normalizar_texto(texto):
    texto = unicodedata.normalize('NFKD', texto or '')
    texto = ''.join(char for char in texto if not unicodedata.combining(char))
    return ' '.join(texto.lower().split())


def _normalizar_nome_conta_depara(texto):
    texto = _normalizar_texto(texto)
    texto = ''.join(char if char.isalnum() else ' ' for char in texto)
    return ' '.join(texto.split())


def _slug_planilha(texto):
    texto = _normalizar_texto(texto).replace(' ', '_')
    return ''.join(char for char in texto if char.isalnum() or char == '_') or 'empresa'


def _obter_anos_disponiveis(campo_data):
    anos = set()

    for modelo in (ContaPagar, ContaReceber):
        anos.update(item.year for item in modelo.objects.dates(campo_data, 'year'))

    if not anos:
        anos.add(date.today().year)

    return sorted(anos, reverse=True)


def _montar_opcoes_periodo(campo_data):
    ano_atual = date.today().year
    anos = range(ano_atual - 5, ano_atual + 6)
    opcoes = [{'valor': 'customizado', 'label': 'Selecionar periodo', 'nivel': 'customizado'}]

    for ano in anos:
        opcoes.append({'valor': f'ano:{ano}', 'label': str(ano), 'nivel': 'ano', 'ano': ano})

        for trimestre, meses in TRIMESTRES.items():
            inicio = MESES_LABELS[meses[0]][:3]
            fim = MESES_LABELS[meses[-1]][:3]
            opcoes.append({
                'valor': f'tri:{ano}:T{trimestre}',
                'label': f'{ano} - T{trimestre} ({inicio} a {fim})',
                'menu_label': f'T{trimestre} ({inicio} a {fim})',
                'nivel': 'trimestre',
                'ano': ano,
                'trimestre': trimestre,
            })

            for mes in meses:
                opcoes.append({
                    'valor': f'mes:{ano}:{mes:02d}',
                    'label': f'{MESES_LABELS[mes]} {ano}',
                    'menu_label': MESES_LABELS[mes],
                    'nivel': 'mes',
                    'ano': ano,
                    'trimestre': trimestre,
                    'mes': mes,
                })

    vistos = set()
    opcoes_unicas = []
    for opcao in opcoes:
        if opcao['valor'] in vistos:
            continue
        vistos.add(opcao['valor'])
        opcoes_unicas.append(opcao)

    return opcoes_unicas


def _primeiro_dia_mes(data_base):
    return date(data_base.year, data_base.month, 1)


def _iterar_meses(data_inicial, data_final):
    referencia = _primeiro_dia_mes(data_inicial)
    limite = _primeiro_dia_mes(data_final)
    meses = []

    while referencia <= limite:
        meses.append({
            'codigo': f'{referencia.year}-{referencia.month:02d}',
            'ano': referencia.year,
            'mes': referencia.month,
            'rotulo': f"{MESES_LABELS[referencia.month][:3]}/{str(referencia.year)[-2:]}",
        })

        if referencia.month == 12:
            referencia = date(referencia.year + 1, 1, 1)
        else:
            referencia = date(referencia.year, referencia.month + 1, 1)

    return meses


def _obter_mes_anterior(mes):
    if mes['mes'] == 1:
        ano_anterior = mes['ano'] - 1
        mes_anterior = 12
    else:
        ano_anterior = mes['ano']
        mes_anterior = mes['mes'] - 1

    return {
        'codigo': f'{ano_anterior}-{mes_anterior:02d}',
        'ano': ano_anterior,
        'mes': mes_anterior,
        'rotulo': MESES_LABELS[mes_anterior],
    }


def _adicionar_mes_referencia_ah(meses_periodo):
    if not meses_periodo:
        return []

    return [_obter_mes_anterior(meses_periodo[0]), *meses_periodo]


def _remover_mes_referencia_ah(dre):
    for linha in dre:
        linha['meses'] = linha['meses'][1:]

    return dre


def _resolver_meses_periodo(periodo_selecionado, data_inicial=None, data_final=None):
    hoje = date.today()
    meses = []

    if periodo_selecionado.startswith('mes:'):
        _, ano_str, mes_str = periodo_selecionado.split(':')
        ano = int(ano_str)
        mes = int(mes_str)
        return [{
            'codigo': f'{ano}-{mes:02d}',
            'ano': ano,
            'mes': mes,
            'rotulo': MESES_LABELS[mes],
        }]

    if periodo_selecionado.startswith('tri:'):
        _, ano_str, trimestre_str = periodo_selecionado.split(':')
        ano = int(ano_str)
        trimestre = int(trimestre_str.replace('T', ''))
        for mes in TRIMESTRES.get(trimestre, (1, 2, 3)):
            meses.append({
                'codigo': f'{ano}-{mes:02d}',
                'ano': ano,
                'mes': mes,
                'rotulo': MESES_LABELS[mes],
            })
        return meses

    if periodo_selecionado == 'customizado' and data_inicial and data_final:
        return _iterar_meses(data_inicial, data_final)

    if periodo_selecionado.startswith('ano:'):
        ano = int(periodo_selecionado.split(':')[1])
    else:
        ano = hoje.year

    for mes in range(1, 13):
        meses.append({
            'codigo': f'{ano}-{mes:02d}',
            'ano': ano,
            'mes': mes,
            'rotulo': MESES_LABELS[mes],
        })

    return meses


def _resolver_periodo_dashboard(request):
    tipo_visualizacao = request.GET.get('tipo_visualizacao') or 'competencia'
    if tipo_visualizacao not in TIPOS_VISUALIZACAO:
        tipo_visualizacao = 'competencia'

    configuracao_visualizacao = TIPOS_VISUALIZACAO[tipo_visualizacao]
    opcoes = _montar_opcoes_periodo(configuracao_visualizacao['campo_data'])
    valores_validos = {opcao['valor'] for opcao in opcoes}
    periodo_selecionado = request.GET.get('periodo') or opcoes[0]['valor']
    data_inicial_str = request.GET.get('data_inicial', '')
    data_final_str = request.GET.get('data_final', '')

    if periodo_selecionado not in valores_validos:
        periodo_selecionado = opcoes[0]['valor']

    data_inicial = None
    data_final = None

    try:
        if data_inicial_str:
            data_inicial = date.fromisoformat(data_inicial_str)
        if data_final_str:
            data_final = date.fromisoformat(data_final_str)
    except ValueError:
        data_inicial = None
        data_final = None

    if periodo_selecionado == 'customizado':
        if not data_inicial or not data_final:
            hoje = date.today()
            data_inicial = date(hoje.year, 1, 1)
            data_final = hoje
        elif data_inicial > data_final:
            data_inicial, data_final = data_final, data_inicial

    meses = _resolver_meses_periodo(periodo_selecionado, data_inicial, data_final)

    if periodo_selecionado == 'customizado':
        legenda = f'{data_inicial.strftime("%d/%m/%Y")} até {data_final.strftime("%d/%m/%Y")}'
    else:
        legenda = next(
            (opcao['label'] for opcao in opcoes if opcao['valor'] == periodo_selecionado),
            periodo_selecionado,
        )

    return {
        'tipo_visualizacao': tipo_visualizacao,
        'tipo_visualizacao_label': configuracao_visualizacao['label'],
        'campo_data_periodo': configuracao_visualizacao['campo_data'],
        'periodo_opcoes': opcoes,
        'periodo_selecionado': periodo_selecionado,
        'meses_dre': meses,
        'periodo_legenda': legenda,
        'mostrar_datas_personalizadas': periodo_selecionado == 'customizado',
        'data_inicial': data_inicial.isoformat() if data_inicial else '',
        'data_final': data_final.isoformat() if data_final else '',
    }


def _mock_valores_conta(conta, indice_base, meses_periodo):
    base = 18000 + (indice_base * 2750)
    valores = {}

    for indice_mes, mes in enumerate(meses_periodo):
        fator = 1 + (indice_mes * 0.06)
        ajuste_sazonal = (mes['mes'] % 3) * 0.015
        valores[mes['codigo']] = float(round(base * (fator + ajuste_sazonal), 2))

    return valores


def _ratear_valores(valores_base, quantidade_partes):
    if quantidade_partes <= 0:
        return []

    pesos = list(range(1, quantidade_partes + 1))
    soma_pesos = sum(pesos)
    rateio = [{codigo: 0.0 for codigo in valores_base.keys()} for _ in range(quantidade_partes)]

    for codigo, valor_total in valores_base.items():
        acumulado = 0.0

        for indice, peso in enumerate(pesos):
            if indice == quantidade_partes - 1:
                valor_parte = round(valor_total - acumulado, 2)
            else:
                valor_parte = round((valor_total * peso) / soma_pesos, 2)
                acumulado += valor_parte

            rateio[indice][codigo] = valor_parte

    return rateio


def _obter_rotulo_cliente_fornecedor(codigo_cliente_fornecedor, nomes_clientes_fornecedores=None):
    if codigo_cliente_fornecedor in (None, ''):
        return 'Sem cliente/fornecedor'

    if nomes_clientes_fornecedores:
        nome = nomes_clientes_fornecedores.get(codigo_cliente_fornecedor)
        if nome:
            return nome

    return f'Cliente/Fornecedor {codigo_cliente_fornecedor}'


def _coletar_movimentos_dre(meses_periodo, campo_data):
    if not meses_periodo:
        return {'positivo': {}, 'negativo': {}}, {'positivo': {}, 'negativo': {}}

    data_inicial = date(meses_periodo[0]['ano'], meses_periodo[0]['mes'], 1)
    data_final = date(meses_periodo[-1]['ano'], meses_periodo[-1]['mes'], 28)
    while True:
        try:
            data_final = date(data_final.year, data_final.month, data_final.day + 1)
        except ValueError:
            break

    meses_validos = {mes['codigo'] for mes in meses_periodo}
    totais_categoria = {'positivo': {}, 'negativo': {}}
    totais_cliente = {'positivo': {}, 'negativo': {}}

    for sinal, modelo in (('positivo', ContaReceber), ('negativo', ContaPagar)):
        filtros = {
            f'{campo_data}__gte': data_inicial,
            f'{campo_data}__lte': data_final,
        }
        queryset = (
            modelo.objects.filter(**filtros)
            .exclude(valor_documento__isnull=True)
            .only(campo_data, 'codigo_categoria', 'codigo_cliente_fornecedor', 'valor_documento')
        )

        for movimento in queryset.iterator():
            data_movimento = getattr(movimento, campo_data, None)
            if not data_movimento:
                continue

            codigo_mes = f'{data_movimento.year}-{data_movimento.month:02d}'
            if codigo_mes not in meses_validos:
                continue

            codigo_categoria = (movimento.codigo_categoria or '').strip()
            if not codigo_categoria:
                continue

            valor = float(movimento.valor_documento or 0)
            cliente_codigo = movimento.codigo_cliente_fornecedor

            valores_categoria = totais_categoria[sinal].setdefault(
                codigo_categoria,
                _inicializar_valores_periodo(meses_periodo),
            )
            valores_categoria[codigo_mes] = round(valores_categoria[codigo_mes] + valor, 2)

            chave_cliente = (codigo_categoria, cliente_codigo)
            valores_cliente = totais_cliente[sinal].setdefault(
                chave_cliente,
                _inicializar_valores_periodo(meses_periodo),
            )
            valores_cliente[codigo_mes] = round(valores_cliente[codigo_mes] + valor, 2)

    return totais_categoria, totais_cliente


def _montar_meses_linha(valores, meses_periodo, referencia_av, multiplicador_exibicao=1):
    meses_linha = []
    valor_anterior = None

    for mes in meses_periodo:
        valor_atual = valores[mes['codigo']]
        valor_exibicao = valor_atual * multiplicador_exibicao
        valor_anterior_exibicao = (
            None if valor_anterior is None else valor_anterior * multiplicador_exibicao
        )
        percentual_ah = None

        if valor_anterior_exibicao not in (None, 0):
            percentual_ah = ((valor_exibicao / valor_anterior_exibicao) - 1) * 100

        meses_linha.append({
            'rotulo': mes['rotulo'],
            'mes_abreviado': MESES_ABREVIADOS[mes['mes']],
            'ano': mes['ano'],
            'realizado': _formatar_moeda(valor_exibicao),
            'av': _formatar_percentual(
                (valor_exibicao / referencia_av[mes['codigo']]) * 100 if referencia_av[mes['codigo']] else 0
            ),
            'ah': '-' if percentual_ah is None else _formatar_percentual(percentual_ah),
            'ah_valor': percentual_ah,
            'ah_direcao': (
                'alta'
                if percentual_ah is not None and percentual_ah > 0
                else 'baixa'
                if percentual_ah is not None and percentual_ah < 0
                else 'neutro'
            ),
        })
        valor_anterior = valor_atual

    return meses_linha


def _construir_linhas_dre(empresa, meses_periodo, campo_data):
    categorias_qs = (
        DeParaCategoriaDRE.objects.filter(empresa=empresa)
        .select_related('categoria')
        .order_by('categoria__codigo', 'categoria__descricao')
    )
    filhos_ativos = (
        ContaDRE.objects.filter(ativo=True)
        .order_by('ordem', 'nome')
        .prefetch_related(Prefetch('categorias_vinculadas', queryset=categorias_qs))
    )
    contas_raiz = (
        ContaDRE.objects.filter(pai__isnull=True, ativo=True)
        .prefetch_related(Prefetch('filhos', queryset=filhos_ativos))
        .order_by('ordem', 'nome')
    )

    dre = []
    acumulado = {mes['codigo']: 0.0 for mes in meses_periodo}
    indice_mock = 1
    totais_categoria, totais_cliente = _coletar_movimentos_dre(meses_periodo, campo_data)
    try:
        nomes_clientes_fornecedores = {
            item.codigo_omie: item.nome_fantasia or item.razao_social
            for item in ClienteFornecedor.objects.all().only('codigo_omie', 'nome_fantasia', 'razao_social')
        }
    except (OperationalError, ProgrammingError):
        nomes_clientes_fornecedores = {}
    possui_movimentos = any(
        any(any(valor for valor in valores.values()) for valores in totais.values())
        for totais in totais_categoria.values()
    )

    referencia_av = None
    nome_ultima_conta_dre = 'resultado do exercicio'

    for conta_pai in contas_raiz:
        filhos = list(conta_pai.filhos.all())
        valores_filhos = []

        if conta_pai.sinal == 'resultado':
            valores_pai = {mes['codigo']: acumulado[mes['codigo']] for mes in meses_periodo}
        elif filhos:
            for filho in filhos:
                categorias_vinculadas = list(filho.categorias_vinculadas.all())

                if possui_movimentos:
                    valores_filho = _inicializar_valores_periodo(meses_periodo)
                    detalhes_categoria = []

                    for item_depara in categorias_vinculadas:
                        categoria = item_depara.categoria
                        valores_categoria = totais_categoria[filho.sinal].get(
                            categoria.codigo,
                            _inicializar_valores_periodo(meses_periodo),
                        ).copy()
                        _somar_valores(valores_filho, valores_categoria)

                        clientes_categoria = []
                        for (codigo_categoria, cliente_codigo), valores_cliente in totais_cliente[filho.sinal].items():
                            if codigo_categoria != categoria.codigo:
                                continue

                            clientes_categoria.append({
                                'codigo': cliente_codigo,
                                'nome': _obter_rotulo_cliente_fornecedor(
                                    cliente_codigo,
                                    nomes_clientes_fornecedores,
                                ),
                                'valores': valores_cliente.copy(),
                            })

                        clientes_categoria.sort(key=lambda item: _normalizar_texto(item['nome']))
                        detalhes_categoria.append((categoria, valores_categoria, clientes_categoria))
                else:
                    valores_filho = _mock_valores_conta(filho, indice_mock, meses_periodo)
                    indice_mock += 1
                    valores_categorias = _ratear_valores(valores_filho, len(categorias_vinculadas))
                    detalhes_categoria = []

                    for indice_categoria, item_depara in enumerate(categorias_vinculadas):
                        detalhes_categoria.append((
                            item_depara.categoria,
                            valores_categorias[indice_categoria],
                            [{
                                'codigo': None,
                                'nome': 'Sem cliente/fornecedor',
                                'valores': valores_categorias[indice_categoria].copy(),
                            }],
                        ))

                valores_filhos.append((filho, valores_filho, detalhes_categoria))

            valores_pai = {
                mes['codigo']: sum(
                    valores_filho[mes['codigo']]
                    for _, valores_filho, _ in valores_filhos
                )
                for mes in meses_periodo
            }
        else:
            valores_pai = (
                _inicializar_valores_periodo(meses_periodo)
                if possui_movimentos
                else _mock_valores_conta(conta_pai, indice_mock, meses_periodo)
            )
            if not possui_movimentos:
                indice_mock += 1

        if referencia_av is None and conta_pai.sinal == 'positivo' and any(valores_pai.values()):
            referencia_av = valores_pai.copy()

        if conta_pai.sinal == 'positivo':
            for mes in meses_periodo:
                acumulado[mes['codigo']] += valores_pai[mes['codigo']]
        elif conta_pai.sinal == 'negativo':
            for mes in meses_periodo:
                acumulado[mes['codigo']] -= valores_pai[mes['codigo']]

        referencia_linha = referencia_av or {mes['codigo']: 1.0 for mes in meses_periodo}
        if conta_pai.sinal == 'positivo' and referencia_av is None and any(valores_pai.values()):
            referencia_linha = {
                mes['codigo']: valores_pai[mes['codigo']] if valores_pai[mes['codigo']] else 1.0
                for mes in meses_periodo
            }
            referencia_av = referencia_linha.copy()

        chave_pai = f'conta-{conta_pai.id}'
        linha_pai = {
            'id': chave_pai,
            'parent_id': '',
            'nivel': 1,
            'sinal': conta_pai.sinal,
            'nome': conta_pai.nome,
            'prefixo': {'positivo': '(+)', 'negativo': '(-)', 'resultado': '(=)'}.get(conta_pai.sinal, ''),
            'meses': _montar_meses_linha(
                valores_pai,
                meses_periodo,
                referencia_linha,
                multiplicador_exibicao=-1 if conta_pai.sinal == 'negativo' else 1,
            ),
            'possui_filhos': bool(valores_filhos),
            'expandido': False,
        }

        dre.append(linha_pai)

        for filho, valores_filho, detalhes_categoria in valores_filhos:
            chave_filho = f'conta-{filho.id}'
            linha_filho = {
                'id': chave_filho,
                'parent_id': chave_pai,
                'nivel': 2,
                'sinal': filho.sinal,
                'nome': filho.nome,
                'prefixo': {'positivo': '(+)', 'negativo': '(-)', 'resultado': '(=)'}.get(filho.sinal, ''),
                'meses': _montar_meses_linha(
                    valores_filho,
                    meses_periodo,
                    referencia_linha,
                    multiplicador_exibicao=-1 if filho.sinal == 'negativo' else 1,
                ),
                'possui_filhos': bool(detalhes_categoria),
                'expandido': False,
            }
            dre.append(linha_filho)

            for categoria, valores_categoria, clientes_categoria in detalhes_categoria:
                chave_categoria = f'categoria-{filho.id}-{categoria.id}'
                linha_categoria = {
                    'id': chave_categoria,
                    'parent_id': chave_filho,
                    'nivel': 3,
                    'sinal': filho.sinal,
                    'nome': categoria.descricao,
                    'prefixo': '',
                    'meses': _montar_meses_linha(
                        valores_categoria,
                        meses_periodo,
                        referencia_linha,
                        multiplicador_exibicao=-1 if filho.sinal == 'negativo' else 1,
                    ),
                    'possui_filhos': bool(clientes_categoria),
                    'expandido': False,
                }
                dre.append(linha_categoria)

                for cliente_categoria in clientes_categoria:
                    linha_cliente = {
                        'id': f"cliente-{filho.id}-{categoria.id}-{cliente_categoria['codigo'] or 'sem'}",
                        'parent_id': chave_categoria,
                        'nivel': 4,
                        'sinal': filho.sinal,
                        'nome': cliente_categoria['nome'],
                        'prefixo': '',
                        'meses': _montar_meses_linha(
                            cliente_categoria['valores'],
                            meses_periodo,
                            referencia_linha,
                            multiplicador_exibicao=-1 if filho.sinal == 'negativo' else 1,
                        ),
                        'possui_filhos': False,
                        'expandido': False,
                    }
                    dre.append(linha_cliente)

        if _normalizar_texto(conta_pai.nome) == nome_ultima_conta_dre:
            break

    if referencia_av is None:
        referencia_av = {mes['codigo']: 1.0 for mes in meses_periodo}

    return dre


def _apurar_valores_contas_dre(empresa, meses_periodo, campo_data, usar_mock=False):
    categorias_qs = (
        DeParaCategoriaDRE.objects.filter(empresa=empresa)
        .select_related('categoria')
        .order_by('categoria__codigo', 'categoria__descricao')
    )
    filhos_ativos = (
        ContaDRE.objects.filter(ativo=True)
        .order_by('ordem', 'nome')
        .prefetch_related(Prefetch('categorias_vinculadas', queryset=categorias_qs))
    )
    contas_raiz = (
        ContaDRE.objects.filter(pai__isnull=True, ativo=True)
        .prefetch_related(Prefetch('filhos', queryset=filhos_ativos))
        .order_by('ordem', 'nome')
    )

    acumulado = {mes['codigo']: 0.0 for mes in meses_periodo}
    indice_mock = 1
    totais_categoria, _ = _coletar_movimentos_dre(meses_periodo, campo_data)
    possui_movimentos = any(
        any(any(valor for valor in valores.values()) for valores in totais.values())
        for totais in totais_categoria.values()
    )
    valores_por_conta = {}
    nome_ultima_conta_dre = 'resultado do exercicio'

    for conta_pai in contas_raiz:
        filhos = list(conta_pai.filhos.all())
        valores_filhos = []

        if conta_pai.sinal == 'resultado':
            valores_pai = {mes['codigo']: acumulado[mes['codigo']] for mes in meses_periodo}
        elif filhos:
            for filho in filhos:
                categorias_vinculadas = list(filho.categorias_vinculadas.all())
                valores_filho = _inicializar_valores_periodo(meses_periodo)

                if possui_movimentos:
                    for item_depara in categorias_vinculadas:
                        valores_categoria = totais_categoria[filho.sinal].get(
                            item_depara.categoria.codigo,
                            _inicializar_valores_periodo(meses_periodo),
                        )
                        _somar_valores(valores_filho, valores_categoria)
                elif usar_mock:
                    valores_filho = _mock_valores_conta(filho, indice_mock, meses_periodo)
                    indice_mock += 1

                valores_filhos.append((filho, valores_filho))
                valores_por_conta[filho.id] = {
                    'conta': filho,
                    'valores': valores_filho.copy(),
                }

            valores_pai = {
                mes['codigo']: sum(
                    valores_filho[mes['codigo']]
                    for _, valores_filho in valores_filhos
                )
                for mes in meses_periodo
            }
        else:
            valores_pai = _inicializar_valores_periodo(meses_periodo)
            if not possui_movimentos and usar_mock:
                valores_pai = _mock_valores_conta(conta_pai, indice_mock, meses_periodo)
                indice_mock += 1

        valores_por_conta[conta_pai.id] = {
            'conta': conta_pai,
            'valores': valores_pai.copy(),
        }

        if conta_pai.sinal == 'positivo':
            for mes in meses_periodo:
                acumulado[mes['codigo']] += valores_pai[mes['codigo']]
        elif conta_pai.sinal == 'negativo':
            for mes in meses_periodo:
                acumulado[mes['codigo']] -= valores_pai[mes['codigo']]

        if _normalizar_texto(conta_pai.nome) == nome_ultima_conta_dre:
            break

    return valores_por_conta


def _mapear_configuracoes_indicadores(empresa):
    try:
        return {
            item.chave: item
            for item in IndicadorConfiguracao.objects.filter(empresa=empresa).select_related('conta_1', 'conta_2', 'conta_3', 'conta_4')
        }
    except (OperationalError, ProgrammingError):
        return {}


def _obter_total_conta(configuracao, campo, valores_por_conta, aplicar_sinal=True):
    conta = getattr(configuracao, campo, None)
    if not conta:
        return None

    dados_conta = valores_por_conta.get(conta.id)
    if not dados_conta:
        return 0.0

    total = round(sum(dados_conta['valores'].values()), 2)
    if aplicar_sinal and conta.sinal == 'negativo':
        total *= -1

    return total


def _montar_indicadores_configurados(empresa):
    configuracoes_salvas = _mapear_configuracoes_indicadores(empresa)
    indicadores = []

    for indicador_base in INDICADORES_CONFIG:
        configuracao = configuracoes_salvas.get(indicador_base['chave'])
        campos = []
        operadores = indicador_base.get('operadores', [])
        for indice_campo, (nome_campo, label_campo) in enumerate(indicador_base['campos']):
            conta = getattr(configuracao, nome_campo, None) if configuracao else None
            campos.append({
                'nome': nome_campo,
                'label': label_campo,
                'valor': conta.id if conta else '',
                'operador_apos': operadores[indice_campo] if indice_campo < len(operadores) else '',
            })

        indicadores.append({
            **indicador_base,
            'campos': campos,
        })

    return indicadores


def _calcular_kpis_dashboard(empresa, meses_periodo, campo_data, usar_mock=False):
    configuracoes_salvas = _mapear_configuracoes_indicadores(empresa)
    valores_por_conta = _apurar_valores_contas_dre(empresa, meses_periodo, campo_data, usar_mock=usar_mock)

    receita_bruta_cfg = configuracoes_salvas.get('receita_bruta')
    receita_bruta_total = _obter_total_conta(receita_bruta_cfg, 'conta_1', valores_por_conta) if receita_bruta_cfg else None

    deducoes_cfg = configuracoes_salvas.get('deducoes_gastos_variaveis')
    deducoes_total = None
    if deducoes_cfg:
        parcela_1 = _obter_total_conta(deducoes_cfg, 'conta_1', valores_por_conta) or 0.0
        parcela_2 = _obter_total_conta(deducoes_cfg, 'conta_2', valores_por_conta) or 0.0
        deducoes_total = round(parcela_1 + parcela_2, 2)

    margem_cfg = configuracoes_salvas.get('margem_contribuicao')
    margem_total = None
    if margem_cfg:
        parcela_base = _obter_total_conta(margem_cfg, 'conta_1', valores_por_conta, aplicar_sinal=False)
        parcela_soma_1 = _obter_total_conta(margem_cfg, 'conta_2', valores_por_conta, aplicar_sinal=False)
        parcela_soma_2 = _obter_total_conta(margem_cfg, 'conta_3', valores_por_conta, aplicar_sinal=False)
        denominador = _obter_total_conta(margem_cfg, 'conta_4', valores_por_conta, aplicar_sinal=False)
        numerador = None
        if (
            parcela_base is not None
            and parcela_soma_1 is not None
            and parcela_soma_2 is not None
        ):
            numerador = parcela_base - (parcela_soma_1 + parcela_soma_2)
        if numerador is not None and denominador:
            margem_total = (numerador / denominador) * 100

    gastos_fixos_cfg = configuracoes_salvas.get('gastos_fixos')
    gastos_fixos_total = _obter_total_conta(gastos_fixos_cfg, 'conta_1', valores_por_conta) if gastos_fixos_cfg else None

    ebit_cfg = configuracoes_salvas.get('ebit')
    ebit_total = None
    if ebit_cfg and receita_bruta_total:
        numerador_ebit = _obter_total_conta(ebit_cfg, 'conta_1', valores_por_conta)
        if numerador_ebit is not None:
            ebit_total = (numerador_ebit / receita_bruta_total) * 100

    return {
        'receita_bruta': _formatar_moeda_resumida(receita_bruta_total) if receita_bruta_total is not None else '-',
        'deducoes_gastos_variaveis': _formatar_moeda_resumida(deducoes_total) if deducoes_total is not None else '-',
        'margem_contribuicao': _formatar_percentual_indicador(margem_total) if margem_total is not None else '-',
        'gastos_fixos': _formatar_moeda_resumida(gastos_fixos_total) if gastos_fixos_total is not None else '-',
        'ebit': _formatar_percentual_indicador(ebit_total) if ebit_total is not None else '-',
    }


def _montar_dados_visao_geral(meses_periodo, campo_data):
    valores_recebidos = _inicializar_valores_periodo(meses_periodo)
    valores_pagos = _inicializar_valores_periodo(meses_periodo)
    data_inicial, data_final = _limites_periodo_meses(meses_periodo)

    for destino, modelo in ((valores_recebidos, ContaReceber), (valores_pagos, ContaPagar)):
        filtros = {
            f'{campo_data}__gte': data_inicial,
            f'{campo_data}__lte': data_final,
        }
        movimentos = (
            modelo.objects.filter(**filtros)
            .exclude(valor_documento__isnull=True)
            .only(campo_data, 'valor_documento')
        )

        for movimento in movimentos.iterator():
            data_movimento = getattr(movimento, campo_data, None)
            if not data_movimento:
                continue

            codigo_mes = f'{data_movimento.year}-{data_movimento.month:02d}'
            if codigo_mes not in destino:
                continue

            destino[codigo_mes] = round(destino[codigo_mes] + float(movimento.valor_documento or 0), 2)

    total_recebido = round(sum(valores_recebidos.values()), 2)
    total_pago = round(sum(valores_pagos.values()), 2)
    resultado = round(total_recebido - total_pago, 2)
    divisor_meses = len(meses_periodo) or 1

    colunas_tempo = []
    colunas_margem = []
    maior_tempo = max(
        [abs(valor) for valor in valores_recebidos.values()]
        + [abs(valor) for valor in valores_pagos.values()]
        + [1]
    )

    for mes in meses_periodo:
        codigo = mes['codigo']
        recebido = valores_recebidos[codigo]
        pago = valores_pagos[codigo]
        resultado_mes = round(recebido - pago, 2)
        margem = (resultado_mes / recebido) * 100 if recebido else 0

        colunas_tempo.append({
            'rotulo': MESES_ABREVIADOS[mes['mes']],
            'ano': mes['ano'],
            'recebimentos': recebido,
            'pagamentos': pago,
            'recebimentos_formatado': _formatar_moeda_resumida(recebido),
            'pagamentos_formatado': _formatar_moeda_resumida(pago),
            'altura_recebimentos': max((abs(recebido) / maior_tempo) * 100, 3) if recebido else 0,
            'altura_pagamentos': max((abs(pago) / maior_tempo) * 100, 3) if pago else 0,
        })

        colunas_margem.append({
            'rotulo': MESES_ABREVIADOS[mes['mes']],
            'ano': mes['ano'],
            'valor': round(margem, 2),
            'percentual': _formatar_percentual_grafico(margem),
            'altura': 0,
        })

    maior_margem = max([abs(coluna['valor']) for coluna in colunas_margem] + [1])
    for coluna in colunas_margem:
        coluna['altura'] = max((abs(coluna['valor']) / maior_margem) * 100, 4) if coluna['valor'] else 0

    return {
        'kpis': {
            'recebimentos': _formatar_moeda_resumida(total_recebido),
            'pagamentos': _formatar_moeda_resumida(total_pago),
            'resultado': _formatar_moeda_resumida(resultado),
            'media_recebimento_mensal': _formatar_moeda_resumida(total_recebido / divisor_meses),
            'media_resultado_mensal': _formatar_moeda_resumida(resultado / divisor_meses),
        },
        'grafico_recebimentos_pagamentos': colunas_tempo,
        'grafico_margem_mensal': colunas_margem,
    }


def _montar_ranking_clientes_fornecedores(meses_periodo, campo_data):
    data_inicial, data_final = _limites_periodo_meses(meses_periodo)
    codigos = set()
    rankings = {}

    for chave, modelo in (('clientes', ContaReceber), ('fornecedores', ContaPagar)):
        filtros = {
            f'{campo_data}__gte': data_inicial,
            f'{campo_data}__lte': data_final,
        }
        itens = list(
            modelo.objects.filter(**filtros)
            .exclude(valor_documento__isnull=True)
            .values('codigo_cliente_fornecedor')
            .annotate(total=Sum('valor_documento'))
            .order_by('-total')[:5]
        )
        rankings[chave] = itens
        codigos.update(item['codigo_cliente_fornecedor'] for item in itens if item['codigo_cliente_fornecedor'])

    nomes = {
        item.codigo_omie: item.nome_fantasia or item.razao_social
        for item in ClienteFornecedor.objects.filter(codigo_omie__in=codigos)
    }

    def normalizar_item(item):
        codigo = item['codigo_cliente_fornecedor']
        total = float(item['total'] or 0)
        return {
            'nome': nomes.get(codigo) or _obter_rotulo_cliente_fornecedor(codigo),
            'valor': total,
            'valor_formatado': _formatar_moeda_resumida(total),
        }

    return {
        'clientes': [normalizar_item(item) for item in rankings['clientes']],
        'fornecedores': [normalizar_item(item) for item in rankings['fornecedores']],
    }


def _montar_colunas_variacao_mensal(linha_referencia):
    valores_validos = [
        abs(mes['ah_valor'])
        for mes in linha_referencia['meses']
        if mes.get('ah_valor') is not None
    ]
    maximo = max(valores_validos + [1])

    colunas = []
    for mes in linha_referencia['meses']:
        valor = mes.get('ah_valor')
        altura = 0 if valor is None else max((abs(valor) / maximo) * 100, 4)
        colunas.append({
            'rotulo': mes['mes_abreviado'],
            'ano': mes['ano'],
            'valor': valor,
            'percentual': '' if valor is None else _formatar_percentual_grafico(valor),
            'altura': f'{altura:.2f}'.replace(',', '.'),
            'direcao': mes['ah_direcao'],
        })

    return colunas


def _montar_grafico_contas_pai_dre(empresa, meses_periodo, campo_data, dre=None):
    valores_por_conta = _apurar_valores_contas_dre(empresa, meses_periodo, campo_data)
    contas_raiz = ContaDRE.objects.filter(pai__isnull=True, ativo=True).order_by('ordem', 'nome')
    contas_grafico = []
    nome_ultima_conta_dre = 'resultado do exercicio'
    linhas_dre_por_id = {
        linha['id']: linha
        for linha in (dre or [])
    }

    for conta in contas_raiz:
        dados_conta = valores_por_conta.get(conta.id)
        if not dados_conta:
            continue

        multiplicador = -1 if conta.sinal == 'negativo' else 1
        valores_periodo = [
            round(float(dados_conta['valores'].get(mes['codigo'], 0.0)) * multiplicador, 2)
            for mes in meses_periodo
        ]
        media_mensal = round(sum(valores_periodo) / len(valores_periodo), 2) if valores_periodo else 0.0

        linha_variacao = linhas_dre_por_id.get(f'conta-{conta.id}')
        colunas_variacao = _montar_colunas_variacao_mensal(linha_variacao) if linha_variacao else []
        anos_variacao = sorted({coluna['ano'] for coluna in colunas_variacao if coluna['ano']})

        contas_grafico.append({
            'id': conta.id,
            'nome': conta.nome,
            'sinal': conta.sinal,
            'media_mensal': _formatar_moeda(media_mensal),
            'variacao_mensal': {
                'titulo': 'Variacao mensal',
                'conta': conta.nome,
                'colunas': colunas_variacao,
                'ano': anos_variacao[0] if len(anos_variacao) == 1 else '',
            },
            'colunas': [
                {
                    'rotulo': MESES_ABREVIADOS[mes['mes']],
                    'valor': valor,
                    'valor_formatado': _formatar_moeda_resumida(valor),
                }
                for mes, valor in zip(meses_periodo, valores_periodo)
            ],
        })

        if _normalizar_texto(conta.nome) == nome_ultima_conta_dre:
            break

    return contas_grafico


def _montar_grafico_variacao_mensal(dre, nome_conta='receita bruta operacional'):
    linha_referencia = next(
        (
            linha for linha in dre
            if _normalizar_texto(linha['nome']) == _normalizar_texto(nome_conta)
        ),
        None,
    )

    if not linha_referencia:
        linha_referencia = next(
            (
                linha for linha in dre
                if linha.get('sinal') == 'positivo' and linha.get('nivel') in (1, 2)
            ),
            None,
        )

    if not linha_referencia:
        return {
            'titulo': 'Variacao mensal',
            'conta': nome_conta,
            'colunas': [],
            'ano': '',
        }

    colunas = _montar_colunas_variacao_mensal(linha_referencia)
    anos = sorted({coluna['ano'] for coluna in colunas if coluna['ano']})

    return {
        'titulo': 'Variacao mensal',
        'conta': linha_referencia['nome'],
        'colunas': colunas,
        'ano': anos[0] if len(anos) == 1 else '',
    }


def _salvar_formula_conta(conta, contas_formula):
    ComposicaoContaDRE.objects.filter(conta_resultado=conta).delete()

    if conta.sinal != 'resultado' or not contas_formula:
        return

    for indice, conta_origem in enumerate(contas_formula, start=1):
        ComposicaoContaDRE.objects.create(
            conta_resultado=conta,
            conta_origem=conta_origem,
            ordem=indice,
        )


def _adicionar_erros_formulario(request, form):
    if form.non_field_errors():
        for erro in form.non_field_errors():
            messages.error(request, erro)

    for campo, erros in form.errors.items():
        if campo == '__all__':
            continue

        label = form.fields[campo].label or campo.replace('_', ' ').capitalize()
        for erro in erros:
            messages.error(request, f'{label}: {erro}')


def dashboard_empresa(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)
    periodo_contexto = _resolver_periodo_dashboard(request)
    dados_visao_geral = _montar_dados_visao_geral(
        periodo_contexto['meses_dre'],
        periodo_contexto['campo_data_periodo'],
    )

    contexto = {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'kpis': dados_visao_geral['kpis'],
        'grafico_recebimentos_pagamentos': dados_visao_geral['grafico_recebimentos_pagamentos'],
        'grafico_margem_mensal': dados_visao_geral['grafico_margem_mensal'],
        'ranking_visao_geral': _montar_ranking_clientes_fornecedores(
            periodo_contexto['meses_dre'],
            periodo_contexto['campo_data_periodo'],
        ),
        **periodo_contexto,
    }
    return render(request, 'dashboard1.html', contexto)


def dashboard_resultado(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)
    periodo_contexto = _resolver_periodo_dashboard(request)
    meses_dre_calculo = _adicionar_mes_referencia_ah(periodo_contexto['meses_dre'])
    dre = _remover_mes_referencia_ah(_construir_linhas_dre(
        empresa,
        meses_dre_calculo,
        periodo_contexto['campo_data_periodo'],
    ))
    grafico_contas_pai_dre = _montar_grafico_contas_pai_dre(
        empresa,
        periodo_contexto['meses_dre'],
        periodo_contexto['campo_data_periodo'],
        dre,
    )
    grafico_variacao_mensal = (
        grafico_contas_pai_dre[0]['variacao_mensal']
        if grafico_contas_pai_dre
        else _montar_grafico_variacao_mensal(dre)
    )

    contexto = {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'kpis': _calcular_kpis_dashboard(
            empresa,
            periodo_contexto['meses_dre'],
            periodo_contexto['campo_data_periodo'],
            usar_mock=True,
        ),
        'dre': dre,
        'grafico_variacao_mensal': grafico_variacao_mensal,
        'grafico_contas_pai_dre': grafico_contas_pai_dre,
        'dre_colspan': 1 + (len(periodo_contexto['meses_dre']) * 3),
        **periodo_contexto,
    }
    return render(request, 'dashboard.html', contexto)


#-------------------------------------------------------------------------------

def parametros_empresa(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)

    contexto = {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'empresa': empresa,
    }
    return render(request, 'parametros.html', contexto)


def parametros_resultado(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)

    contexto = {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'empresa': empresa,
    }
    return render(request, 'parametros_resultado.html', contexto)


def parametros_gerais(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)

    modal_mensagem = None
    modal_tipo = None

    if request.method == 'POST':
        acao = request.POST.get('acao')

        if acao == 'salvar_credenciais':
            form_parametros = ParametroEmpresaForm(request.POST, instance=empresa)
            if form_parametros.is_valid():
                empresa = form_parametros.save()
                modal_mensagem = 'Credenciais da Omie salvas com sucesso.'
                modal_tipo = 'success'
            else:
                modal_mensagem = 'Erro ao salvar as credenciais da Omie.'
                modal_tipo = 'error'
        elif acao == 'sincronizar_omie':
            form_parametros = ParametroEmpresaForm(instance=empresa)

            if not empresa.app_key_omie or not empresa.app_secret_omie:
                modal_mensagem = 'Salve a App Key e a App Secret antes de atualizar os cadastros e lancamentos.'
                modal_tipo = 'error'
            else:
                try:
                    totais = sincronizar_dados_omie(
                        empresa.app_key_omie,
                        empresa.app_secret_omie,
                    )
                    modal_mensagem = _montar_mensagem_sincronizacao_concluida(totais)
                    modal_tipo = 'success'
                except Exception as e:
                    modal_mensagem = f'Erro ao sincronizar dados da Omie: {e}'
                    modal_tipo = 'error'
        else:
            form_parametros = ParametroEmpresaForm(instance=empresa)

    else:
        form_parametros = ParametroEmpresaForm(instance=empresa)

    contexto = {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'empresa': empresa,
        'form_parametros': form_parametros,
        'modal_mensagem': modal_mensagem,
        'modal_tipo': modal_tipo,
    }
    return render(request, 'parametros_gerais.html', contexto)


def iniciar_sincronizacao_omie(request, slug):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'mensagem': 'Metodo nao permitido.'}, status=405)

    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)
    if not empresa.app_key_omie or not empresa.app_secret_omie:
        return JsonResponse({
            'ok': False,
            'mensagem': 'Salve a App Key e a App Secret antes de atualizar os cadastros e lancamentos.',
        }, status=400)

    task_id = _iniciar_sincronizacao_omie(empresa)
    return JsonResponse({
        'ok': True,
        'task_id': task_id,
        'mensagem': 'Sincronizacao iniciada com sucesso.',
    })


def status_sincronizacao_omie(request, slug, task_id):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)
    status = _obter_status_sincronizacao(task_id)

    if not status or status.get('empresa_id') != empresa.id:
        return JsonResponse({'ok': False, 'mensagem': 'Sincronizacao nao encontrada.'}, status=404)

    return JsonResponse({'ok': True, **status})


#-------------------------------------------------------------------------------

def contas_dre(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)

    conta_edicao = None
    conta_id = request.GET.get('editar')

    if conta_id:
        conta_edicao = get_object_or_404(ContaDRE, id=conta_id)

    if request.method == 'POST':
        conta_id_post = request.POST.get('conta_id')

        if conta_id_post:
            conta_edicao = get_object_or_404(ContaDRE, id=conta_id_post)
            form = ContaDREForm(request.POST, instance=conta_edicao)
            mensagem_sucesso = 'Conta DRE atualizada com sucesso.'
        else:
            form = ContaDREForm(request.POST)
            mensagem_sucesso = 'Conta DRE cadastrada com sucesso.'

        if form.is_valid():
            conta = form.save()
            contas_formula = form.cleaned_data.get('contas_formula')
            _salvar_formula_conta(conta, contas_formula)

            messages.success(request, mensagem_sucesso)
            return redirect('contas_dre', slug=slug)

        _adicionar_erros_formulario(request, form)
    else:
        if conta_edicao:
            form = ContaDREForm(instance=conta_edicao)
        else:
            form = ContaDREForm()

    filhos_ativos = ContaDRE.objects.filter(ativo=True).order_by('ordem', 'nome')
    contas_raiz = (
        ContaDRE.objects.filter(
            pai__isnull=True,
            ativo=True,
        )
        .prefetch_related(Prefetch('filhos', queryset=filhos_ativos))
        .order_by('ordem', 'nome')
    )

    contexto = {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'form': form,
        'contas_raiz': contas_raiz,
        'conta_edicao': conta_edicao,
        'limite_niveis_dre': 2,
    }
    return render(request, 'contas_dre.html', contexto)


#-------------------------------------------------------------------------------

def indicadores(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)

    contas_kpi = ContaDRE.objects.filter(
        ativo=True,
        kpi=True,
    ).order_by('ordem', 'nome')

    indicadores_config = [
        {'titulo': 'Receita Bruta', 'campos': ['receita_bruta'], 'operador': None},
        {
            'titulo': 'Deducoes e Gastos Variáveis',
            'campos': ['deducoes_gastos_variaveis_1', 'deducoes_gastos_variaveis_2'],
            'operador': '+',
        },
        {
            'titulo': 'Margem de Contribuição',
            'campos': ['margem_contribuicao_1', 'margem_contribuicao_2'],
            'operador': '/',
        },
        {'titulo': 'Gastos Fixos', 'campos': ['gastos_fixos'], 'operador': None},
        {'titulo': 'EBIT', 'campos': ['ebit'], 'operador': None},
    ]

    if request.method == 'POST':
        try:
            for indicador_base in INDICADORES_CONFIG:
                defaults = {}
                possui_valor = False

                for nome_campo, _ in indicador_base['campos']:
                    conta_id = request.POST.get(f"{indicador_base['chave']}__{nome_campo}")
                    conta = contas_kpi.filter(id=conta_id).first() if conta_id else None
                    defaults[nome_campo] = conta
                    possui_valor = possui_valor or bool(conta)

                if possui_valor:
                    IndicadorConfiguracao.objects.update_or_create(
                        empresa=empresa,
                        chave=indicador_base['chave'],
                        defaults=defaults,
                    )
                else:
                    IndicadorConfiguracao.objects.filter(
                        empresa=empresa,
                        chave=indicador_base['chave'],
                    ).delete()

            messages.success(request, 'Indicadores atualizados com sucesso.')
        except (OperationalError, ProgrammingError):
            messages.warning(
                request,
                'A configuracao dos indicadores ainda nao esta disponivel. Execute as migrations para habilitar o salvamento.',
            )
        return redirect('indicadores', slug=slug)

    contexto = {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'contas_kpi': contas_kpi,
        'indicadores_config': _montar_indicadores_configurados(empresa),
    }
    return render(request, 'indicadores.html', contexto)


#-------------------------------------------------------------------------------

def excluir_conta_dre(request, slug, id):
    get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)
    conta = get_object_or_404(ContaDRE, id=id)

    if request.method == 'POST':
        if conta.filhos.exists():
            messages.error(request, 'Não é possível excluir uma conta que possui subcontas.')
        else:
            conta.delete()
            messages.success(request, 'Conta DRE excluída com sucesso.')

    return redirect('contas_dre', slug=slug)


#-------------------------------------------------------------------------------

@require_POST
def reordenar_contas_dre(request, slug):
    get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'ok': False, 'mensagem': 'Dados invalidos.'}, status=400)

    conta_ids = payload.get('contas') or []
    if not isinstance(conta_ids, list) or not conta_ids:
        return JsonResponse({'ok': False, 'mensagem': 'Lista de contas invalida.'}, status=400)

    try:
        conta_ids = [int(conta_id) for conta_id in conta_ids]
    except (TypeError, ValueError):
        return JsonResponse({'ok': False, 'mensagem': 'Lista de contas invalida.'}, status=400)

    contas = list(ContaDRE.objects.filter(id__in=conta_ids, ativo=True).select_related('pai'))
    if len(contas) != len(set(conta_ids)):
        return JsonResponse({'ok': False, 'mensagem': 'Conta nao encontrada.'}, status=404)

    contas_por_id = {conta.id: conta for conta in contas}
    primeira_conta = contas_por_id[conta_ids[0]]
    pai_id = primeira_conta.pai_id

    if any(conta.pai_id != pai_id for conta in contas):
        return JsonResponse({'ok': False, 'mensagem': 'Reordene apenas contas do mesmo nivel.'}, status=400)

    with transaction.atomic():
        for indice, conta_id in enumerate(conta_ids, start=1):
            ContaDRE.objects.filter(id=conta_id).update(ordem=indice)

    return JsonResponse({'ok': True})


#-------------------------------------------------------------------------------

def _listar_categorias_depara():
    categorias_qs = Categoria.objects.filter(
        Q(conta_inativa__isnull=True) | Q(conta_inativa='') | Q(conta_inativa='N')
    ).order_by('codigo', 'descricao')

    categorias = []
    for categoria in categorias_qs:
        codigo = (categoria.codigo or '').strip()
        partes = codigo.split('.')

        eh_pai = len(partes) == 2

        prefixo = codigo[:1]
        if prefixo == '1':
            tipo = 'receita'
        elif prefixo == '2':
            tipo = 'despesa'
        else:
            tipo = 'neutra'

        categorias.append({
            'obj': categoria,
            'eh_pai': eh_pai,
            'tipo': tipo,
        })

    return categorias


def _obter_contas_dre_depara():
    return ContaDRE.objects.filter(
        ativo=True,
        nivel=2,
        sinal__in=['positivo', 'negativo']
    ).order_by('ordem', 'nome')


def _montar_contexto_depara(empresa, slug, modal_mensagem=None, modal_tipo=None):
    categorias = _listar_categorias_depara()
    contas_dre = _obter_contas_dre_depara()
    mapeamentos = {
        item.categoria_id: item.conta_dre_id
        for item in DeParaCategoriaDRE.objects.filter(empresa=empresa)
    }

    return {
        'slug': slug,
        'empresa_nome': empresa.nome_empresa,
        'empresa': empresa,
        'categorias': categorias,
        'contas_dre': contas_dre,
        'mapeamentos': mapeamentos,
        'modal_mensagem': modal_mensagem,
        'modal_tipo': modal_tipo,
    }


def exportar_planilha_depara(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)
    categorias = _listar_categorias_depara()
    mapeamentos = {
        item.categoria_id: item.conta_dre.nome if item.conta_dre else ''
        for item in DeParaCategoriaDRE.objects.filter(empresa=empresa).select_related('conta_dre')
    }

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'DePara DRE'
    worksheet.append(['Codigo da Categoria', 'Nome da Categoria', 'Conta DRE'])

    for item in categorias:
        categoria = item['obj']
        if item['eh_pai']:
            conta_dre = 'Conta totalizadora'
        else:
            conta_dre = mapeamentos.get(categoria.id, '')

        worksheet.append([
            categoria.codigo,
            categoria.descricao,
            conta_dre,
        ])

        if item['eh_pai']:
            linha_atual = worksheet.max_row
            for coluna in range(1, 4):
                worksheet.cell(row=linha_atual, column=coluna).font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nome_arquivo = f"depara_categorias_dre_{_slug_planilha(empresa.nome_empresa)}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    workbook.save(response)
    return response


def depara_categoria_dre(request, slug):
    empresa = get_object_or_404(ParametroEmpresa, slug_empresa__iexact=slug)
    categorias = _listar_categorias_depara()
    contas_dre = _obter_contas_dre_depara()

    if request.method == 'POST':
        if request.POST.get('acao') == 'importar_planilha':
            arquivo = request.FILES.get('planilha_depara')
            if not arquivo:
                contexto = _montar_contexto_depara(
                    empresa,
                    slug,
                    modal_mensagem='Selecione uma planilha .xlsx para importar.',
                    modal_tipo='error',
                )
                return render(request, 'depara_categoria_dre.html', contexto)

            try:
                workbook = load_workbook(arquivo, data_only=True)
                worksheet = workbook.active
            except Exception:
                contexto = _montar_contexto_depara(
                    empresa,
                    slug,
                    modal_mensagem='Nao foi possivel ler a planilha. Envie um arquivo .xlsx valido.',
                    modal_tipo='error',
                )
                return render(request, 'depara_categoria_dre.html', contexto)

            categorias_por_codigo = {
                item['obj'].codigo: item
                for item in categorias
            }
            contas_por_nome = {
                conta.nome.strip(): conta
                for conta in contas_dre
            }
            contas_por_nome_normalizado = {
                _normalizar_nome_conta_depara(conta.nome): conta
                for conta in contas_dre
            }
            atualizacoes = []
            erros = []

            for linha_idx, linha in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                codigo = str((linha[0] if len(linha) > 0 and linha[0] is not None else '')).strip()
                conta_nome = str((linha[2] if len(linha) > 2 and linha[2] is not None else '')).strip()

                if not codigo:
                    continue

                item_categoria = categorias_por_codigo.get(codigo)
                if not item_categoria:
                    continue

                if item_categoria['eh_pai']:
                    continue

                if not conta_nome or conta_nome == 'Conta totalizadora':
                    continue

                conta_dre = contas_por_nome.get(conta_nome)
                if not conta_dre:
                    conta_dre = contas_por_nome_normalizado.get(
                        _normalizar_nome_conta_depara(conta_nome)
                    )
                if not conta_dre:
                    erros.append(
                        f'Linha {linha_idx}: a conta DRE "{conta_nome}" nao foi encontrada para a categoria {codigo}.'
                    )
                    continue

                atualizacoes.append((item_categoria['obj'], conta_dre))

            if erros:
                contexto = _montar_contexto_depara(
                    empresa,
                    slug,
                    modal_mensagem='A planilha nao foi importada por motivos de preenchimento: ' + ' '.join(erros),
                    modal_tipo='error',
                )
                return render(request, 'depara_categoria_dre.html', contexto)

            for categoria, conta_dre in atualizacoes:
                DeParaCategoriaDRE.objects.update_or_create(
                    empresa=empresa,
                    categoria=categoria,
                    defaults={'conta_dre': conta_dre}
                )

            contexto = _montar_contexto_depara(
                empresa,
                slug,
                modal_mensagem='Planilha importada com sucesso.',
                modal_tipo='success',
            )
            return render(request, 'depara_categoria_dre.html', contexto)

        for item in categorias:
            categoria = item['obj']

            if item['eh_pai']:
                continue

            campo = f'conta_categoria_{categoria.id}'
            conta_dre_id = request.POST.get(campo)

            if conta_dre_id:
                conta_dre = ContaDRE.objects.filter(id=conta_dre_id).first()
                DeParaCategoriaDRE.objects.update_or_create(
                    empresa=empresa,
                    categoria=categoria,
                    defaults={'conta_dre': conta_dre}
                )
            else:
                DeParaCategoriaDRE.objects.filter(
                    empresa=empresa,
                    categoria=categoria
                ).delete()

        messages.success(request, 'De/Para salvo com sucesso.')
        return redirect('depara_categoria_dre', slug=slug)
    contexto = _montar_contexto_depara(empresa, slug)
    return render(request, 'depara_categoria_dre.html', contexto)


#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

def mover_conta_dre(request, slug, id, direcao):
    conta = get_object_or_404(ContaDRE, id=id)

    if conta.nivel not in (1, 2):
        return redirect('contas_dre', slug=slug)

    irmaos = ContaDRE.objects.filter(
        pai=conta.pai,
        ativo=True
    ).order_by('ordem', 'id')

    irmaos_list = list(irmaos)
    index = irmaos_list.index(conta)

    if direcao == 'cima' and index > 0:
        conta_anterior = irmaos_list[index - 1]
        conta.ordem, conta_anterior.ordem = conta_anterior.ordem, conta.ordem
        conta.save()
        conta_anterior.save()

    elif direcao == 'baixo' and index < len(irmaos_list) - 1:
        conta_proxima = irmaos_list[index + 1]
        conta.ordem, conta_proxima.ordem = conta_proxima.ordem, conta.ordem
        conta.save()
        conta_proxima.save()

    return redirect('contas_dre', slug=slug)
